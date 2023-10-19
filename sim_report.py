
class SimReport:
    def __init__(self,unit):
        import openpyxl
       
        self.workbook = openpyxl.Workbook()
        self.unit=unit
        self.sheet = self.workbook.active
        self.head_sheet = self.workbook.create_sheet(title="Header")
        self.inj_sheet = self.workbook.create_sheet(title="INJ_WELLS")
        self.P_sheet=self.workbook.create_sheet(title="Pressure")
        self.inplace_sheet=self.workbook.create_sheet(title="Initial in-place")
        self.head_column_names = ["Rlat","Rlong","i","j","x","y","API","WellName","FluidType","WellStatus","BHLatitude","BHLongitude","CurrentOperator","OriginalOperator","SpudDate","CompletionDate","MeasuredDepth","TrueVerticalDepth"]
        self.sheet.title = "PROD_WELLS"

        if self.unit=="FIELD": 
                self.sheet_column_names = ["API", "WellName", "ReportDate", "Days", "WellOil (STB)", "WellGas (MSCF)", "WellWater (STB)"]
                self.inj_column_names = ["API", "WellName", "ReportDate", "Days",  "MonthlyWater (bbl)","MonthlyGas (mcf)"]
                self.P_column_names = ["Date","Pressure(Psi)"]  
                self.inplace_column_names = ["OIL (STB)","FREE GAS (Mscf)","WATER (STB)","PORE-VOLUME(RB)"]  
        elif self.unit=="METRIC": 
                self.sheet_column_names = ["API", "WellName", "ReportDate", "Days", "WellOil (SM3)", "WellGas (SM3)", "WellWater (SM3)"]
                self.inj_column_names = ["API", "WellName", "ReportDate", "Days",  "MonthlyWater (SM3)","MonthlyGas (SM3)"]
                self.P_column_names = ["Date","Pressure(BARSA)"]  
                self.inplace_column_names = ["OIL (STB)","FREE GAS (SM3)","WATER (SM3)","PORE-VOLUME(RM3)"]  
        else:
            raise ValueError("Invalid unit. Please enter a valid unit, METRIC or FIELD.")
                
    def input_colum(self,column_names,sheet_name):
            for col_num, column_name in enumerate(column_names, start=1):
                sheet_name.cell(row=1, column=col_num).value = column_name
            sheet_name.sheet_format.defaultColWidth  = 15
    def sql_r(self,case_num,sql_t=None):
        import rips
        import pyodbc
        resinsight = rips.Instance.find()
        if resinsight is not None:
                project = resinsight.project

                summary_case = project.summary_case(case_num)

                if summary_case is None:
                    print("No summary case found")
                    exit()
                else:
                    cases = resinsight.project.cases()
                    if sql_t=='ACTIVE':
                        server = input('enter the server name:')
                        database = input('enter the database name:')
                        username = input('enter your username:')
                        password = input('enter your password :')

                        # Establish a connection
                        self.connection_string = f'DRIVER={{SQL Server}};SERVER={server};DATABASE={database};UID={username};PWD={password}'
                        self.connection = pyodbc.connect(self.connection_string)
                        self.cursor = self.connection.cursor()
                        self.cursor.execute(f"SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{cases[case_num-1].name}_PROD_WELL'")
                        if self.cursor.fetchone():
                                print("Table 'PROD_WELL' already exists.")
                        else:


                        # Define the SQL table creation query
                            self.table_creation_query = f"""
                                CREATE TABLE  {cases[case_num-1].name}_PROD_WELL (
                                    WellName NVARCHAR(255),
                                    ReportDate DATE,
                                    Days REAL,
                                    WellOil REAL,
                                    WellGas REAL,
                                    WellWater REAL
                                )
                            """

                            
                            # Commit the changes
                            self.cursor.execute(self.table_creation_query)
                        self.cursor.execute(f"SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{cases[case_num-1].name}_HEADER'")
                        if self.cursor.fetchone():
                                print("Table 'HEADER' already exists.")
                        else:


                        # Define the SQL table creation query
                                                                    

                            self.header_creation_query = f"""
                                CREATE TABLE  {cases[case_num-1].name}_HEADER (
                                    Rlat REAL,
                                    Rlong REAL,
                                    i REAL,
                                    j REAL,
                                    x REAL,
                                    y REAL,
                                    API NVARCHAR(255),
                                    WellName NVARCHAR(255),
                                    FluidType NVARCHAR(255),
                                    WellStatus NVARCHAR(255),
                                    BHLatitude REAL,
                                    BHLongitude REAL,
                                    CurrentOperator NVARCHAR(255),
                                    OriginalOperator NVARCHAR(255),
                                    SpudDate DATE,
                                    CompletionDate DATE,
                                    MeasuredDepth REAL,
                                    TrueVerticalDepth REAL,
                                )
                            """
                                                

                            # Commit the changes
                            self.cursor.execute(self.header_creation_query)
                        self.cursor.execute(f"SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{cases[case_num-1].name}_INJ_WELLS'")
                        if self.cursor.fetchone():
                                print("Table 'INJ_WELLS' already exists.")
                        else:

                            self.INJ_WELLS_creation_query = f"""
                                CREATE TABLE  {cases[case_num-1].name}_INJ_WELLS (
                                    API NVARCHAR(255),
                                    WellName NVARCHAR(255),
                                    ReportDate DATE,
                                    Days REAL,
                                    MonthlyWater REAL,
                                    MonthlyGasREAL REAL
                                )
                            """
                            
                            # Commit the changes
                            self.cursor.execute(self.INJ_WELLS_creation_query)
                            
                            
                        self.cursor.execute(f"SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{cases[case_num-1].name}_PRESSURE'")
                        if self.cursor.fetchone():
                                print("Table 'PRESSURE' already exists.")
                        else:


                            self.PRESSURE_creation_query = f"""
                                CREATE TABLE  {cases[case_num-1].name}_PRESSURE (
                                    Pressure REAL,
                                    Date DATE,
                                )
                            """
                            
                            # Commit the changes
                            self.cursor.execute(self.PRESSURE_creation_query)                


                        self.cursor.execute(f"SELECT 1 FROM INFORMATION_SCHEMA.TABLES WHERE TABLE_NAME = '{cases[case_num-1].name}_Initial_in_place'")
                        if self.cursor.fetchone():
                                print("Table 'Initial_in_place' already exists.")
                        else:



                            self.Initial_in_place_creation_query = f"""
                                CREATE TABLE  {cases[case_num-1].name}_Initial_in_place (
                                    OIL REAL,
                                    FREE_GAS REAL,
                                    WATER REAL,
                                    PORE_VOLUME REAL,
                                )
                            """
                            
                            # Commit the changes
                            self.cursor.execute(self.Initial_in_place_creation_query)                 
                            
                        return self.connection.commit()
                    
                    
                    
                
    def sheet_report(self,case_num,Rlat,Rlong,saving_path=None,MeasuredDepth=None,TrueVerticalDepth=None,sql_t=None,CurrentOperator=None,OriginalOperator=None):
        if sql_t=="ACTIVE":
            self.sql_r(case_num=case_num,sql_t="ACTIVE")
        import rips
        import numpy as np
        import time
        self.input_colum( self.sheet_column_names,self.sheet)
        self.input_colum( self.head_column_names,self.head_sheet)
        self.input_colum( self.inj_column_names,self.inj_sheet)
        self.input_colum( self.P_column_names,self.P_sheet)
        self.input_colum( self.inplace_column_names,self.inplace_sheet)
        
        n_well=[]
    
        
        n_time=[]
        n_step=[]
        num =2    
        num1=2
        num2=2
        num3=2
        num4=2
        num5=2
        num6=2
        num7=2
        num10=0
        resinsight = rips.Instance.find()
        if resinsight is not None:
                project = resinsight.project

                summary_case = project.summary_case(case_num)

                if summary_case is None:
                    print("No summary case found")
                    exit()
                else:
                    cases = resinsight.project.cases()
                    OOIP=np.array(summary_case.resample_values("FOIP").values)
                    if len(OOIP)>0: 
 
                        self.inplace_sheet.cell(row=2, column=1).value =OOIP[0]

                    OGIP=np.array(summary_case.resample_values("FGIP").values)
                    if len(OGIP)>0:
                        self.inplace_sheet.cell(row=2, column=2).value =OGIP[0]

                    OWIP=np.array(summary_case.resample_values("FWIP").values)
                    if len(OWIP)>0:
                        self.inplace_sheet.cell(row=2, column=3).value =OWIP[0]
                    
                    porv_results = cases[case_num-1].active_cell_property("STATIC_NATIVE", "PORV", 0)
                    pov=0
                    for i in range(len(porv_results)):
                            pov=porv_results[i]+pov
                    
                    self.inplace_sheet.cell(row=2, column=4).value =pov
                    FRP=np.array(summary_case.resample_values("FPR").values)
                    
                    for p, P in enumerate(FRP, start=2):
                            r5 = p  
                            self.P_sheet.cell(row=r5, column=2).value =P
                            summary_data_sampled = summary_case.resample_values("FPR")
                            t = summary_data_sampled.time_steps[num10]  # Get the corresponding time step    
                            if sql_t=='ACTIVE':
                             self.cursor.execute(f'INSERT INTO {cases[case_num-1].name}_PRESSURE (Date,Pressure) VALUES (?,?) ', (time.strftime("%d %b %Y", time.gmtime(t)),P))
                             self.connection.commit()
                            self.P_sheet.cell(row=r5, column=1).value = time.strftime("%d %b %Y", time.gmtime(t))
                            num10+=1

 
                    sim_wells = cases[case_num-1].simulation_wells() 
                    steps = summary_case.summary_vector_values('TIME').values
                    step = [steps[0]]

                    st =  cases[case_num-1].time_steps()

                    for o in range(len(steps)-1): 
                                    step.append(steps[o+1] - steps[o])

                    n=len(st)-1

                    for v,sim_well in enumerate(sim_wells, start=num7):
                        r8=v
                        self.head_sheet.cell(row=r8, column=7).value = sim_well.name
                        self.head_sheet.cell(row=r8, column=8).value = sim_well.name
                        cell=sim_well.cells(n)[case_num-1]
                        i=cell.ijk.i + 1
                        j=cell.ijk.j + 1
                        grid = cases[case_num-1].grids()[cell.grid_index]
                        dimensions = grid.dimensions()
                        cell_index = (dimensions.i * dimensions.j * cell.ijk.k+ dimensions.i * cell.ijk.j+ cell.ijk.i)
                        cell_centers = grid.cell_centers()
                        cell_center = cell_centers[cell_index]
                    

                        self.head_sheet.cell(row=r8, column=3).value=i
                        self.head_sheet.cell(row=r8, column=4).value=j

                        
                        self.head_sheet.cell(row=r8, column=18).value =TrueVerticalDepth

                        
                        self.head_sheet.cell(row=r8, column=17).value =MeasuredDepth


                        self.head_sheet.cell(row=2, column=1).value =Rlat
                        self.head_sheet.cell(row=2, column=2).value =Rlong
                     
                        dx=cell_center.x
                        dy=cell_center.y
                    

                        self.head_sheet.cell(row=r8, column=5).value =dx
                        self.head_sheet.cell(row=r8, column=6).value =dy
                        
                        if self.unit=="FIELD":
                            
                            bh=(Rlat*364543.98+dy)*1/364543.98
                            self.head_sheet.cell(row=r8, column=11).value =bh
                            rh=(Rlong*np.cos(Rlat)*364543.98+dx)*(1/(np.cos(Rlat)*364543.98))
                            self.head_sheet.cell(row=r8, column=12).value =rh
                            
                        elif self.unit=="METRIC":
                            bh=(Rlat*364543.98+dy*3.28084)*1/364543.98
                            self.head_sheet.cell(row=r8, column=11).value =bh
                            rh=(Rlong*np.cos(Rlat)*364543.98+dx*3.28084)*(1/(np.cos(Rlat)*364543.98))
                            self.head_sheet.cell(row=r8, column=12).value =rh
                            

                        c=0
                        c1=0
                        c2=0
                        WOPT = np.array(summary_case.resample_values(f"WOPT:{sim_well.name}").values)
                        WGIT=np.array(summary_case.resample_values(f"WGIT:{sim_well.name}").values)
                        WWIT=np.array(summary_case.resample_values(f"WWIT:{sim_well.name}").values)
           

                        if len(WOPT)>0:
                                for v1 in WOPT:
                                            if v1 != 0:
                                                break
                                            c += 1

                                summary_data_sampled = summary_case.resample_values(f"WOPT:{sim_well.name}")

                                t0=summary_data_sampled.time_steps
                                t0=t0[c:] 
                                
                                if len(t0)>0:
                                    
                                                tt=time.strftime("%d %b %Y", time.gmtime(t0[0]))
                                                self.head_sheet.cell(row=r8, column=15).value=tt
                                                self.head_sheet.cell(row=r8, column=16).value=tt
                                                self.head_sheet.cell(row=r8, column=9).value="OIL"
                                                self.head_sheet.cell(row=r8, column=10).value="OPEN"
                                                stat="OPEN"
                                                fluid="OIL"
                                                
                        elif len(WGIT)>0 and  len(WWIT)==0 :
                            
                            for v2 in WGIT:
                                        if v2 != 0:
                                            break
                                        c1 += 1

                            summary_data_sampled = summary_case.resample_values(f"WGIT:{sim_well.name}")

                            t1=summary_data_sampled.time_steps
                            t1=t1[c1:] 
                            if len(t1)>0:
                                                tt=time.strftime("%d %b %Y", time.gmtime(t1[0]))
                                                self.head_sheet.cell(row=r8, column=15).value=tt
                                                self.head_sheet.cell(row=r8, column=16).value=tt
                                                self.head_sheet.cell(row=r8, column=9).value="GAS"
                                                self.head_sheet.cell(row=r8, column=10).value="OPEN"
                                                stat="OPEN"
                                                fluid="GAS"
                                                
                        elif len(WWIT)>0 and len(WGIT)==0:
                            for v3 in WWIT:
                                        if v3 != 0:
                                            break
                                        c2 += 1

                            summary_data_sampled = summary_case.resample_values(f"WWIT:{sim_well.name}")

                            t2=summary_data_sampled.time_steps
                            t2=t2[c2:] 
                            if len(t2)>0:
                                                tt=time.strftime("%d %b %Y", time.gmtime(t2[0]))
                                                self.head_sheet.cell(row=r8, column=15).value=tt
                                                self.head_sheet.cell(row=r8, column=16).value=tt
                                                self.head_sheet.cell(row=r8, column=9).value="WATER"
                                                self.head_sheet.cell(row=r8, column=10).value="OPEN"
                                                stat="OPEN"
                                                fluid="WATER"
                                                
                        elif len(WWIT)>0 and len(WGIT)>0 :
                            for v3 in WWIT:
                                        if v3 != 0:
                                            break
                                        c2 += 1

                            summary_data_sampled = summary_case.resample_values(f"WWIT:{sim_well.name}")

                            t2=summary_data_sampled.time_steps
                            t2=t2[c2:] 
                            
                            if len(t2)>0:
                                                tt=time.strftime("%d %b %Y", time.gmtime(t2[0]))
                                                self.head_sheet.cell(row=r8, column=15).value=tt
                                                self.head_sheet.cell(row=r8, column=16).value=tt
                                                self.head_sheet.cell(row=r8, column=9).value="WATER and GAS"
                                                self.head_sheet.cell(row=r8, column=10).value="OPEN"
                                                stat="OPEN"
                                                fluid="WATER and GAS"
                    
                        self.head_sheet.cell(row=r8, column=13).value=CurrentOperator
                        self.head_sheet.cell(row=r8, column=14).value=OriginalOperator
                        if sql_t=='ACTIVE':
                         insert_head = f"INSERT INTO {cases[case_num-1].name}_HEADER (Rlat,Rlong,i,j,x,y,API,WellName,FluidType,WellStatus,BHLatitude,	BHLongitude,CurrentOperator,OriginalOperator,SpudDate,CompletionDate,MeasuredDepth,TrueVerticalDepth) VALUES (?,?,?, ?, ?, ?,?,?,?,?,?,?,?,?,?,?,?,?)"
                         
                         values_to_head = (Rlat,Rlong,i,j,dx,dy,sim_well.name,sim_well.name,fluid,stat,bh,rh,CurrentOperator,OriginalOperator,tt,tt,MeasuredDepth,TrueVerticalDepth)


                         self.cursor.execute(insert_head, values_to_head)
                         self.connection.commit()

                    num7 = num7 + len(sim_wells)
                    n_wgi=[]
                    n_wwi=[]
                    for sim_well in sim_wells:
                            
                            WOPT = np.array(summary_case.resample_values(f"WOPT:{sim_well.name}").values)
                            WGPT = np.array(summary_case.resample_values(f"WGPT:{sim_well.name}").values)
                            
                            
                            WWPT = np.array(summary_case.resample_values(f"WWPT:{sim_well.name}").values)


                            if len(WGPT) > 0:  # Check if WOPT array is not empty
                                WGP = [WGPT[0]]
                                for i2 in range(len(WGPT)-1): 

                                    WGP.append(WGPT[i2+1] - WGPT[i2])
                                WGPar=np.array(WGP)


                                for i2, value in enumerate(WGP, start=num3):
                                    row2 = i2
                                    self.sheet.cell(row=row2, column=6).value = value
                                     

                                num3 = num3 + len(WGP)
                            else:
                                pass

                            if len(WWPT) > 0:  # Check if WOPT array is not empty
                                WWP = [WWPT[0]]
                                for i3 in range(len(WWPT)-1): 

                                    WWP.append(WWPT[i3+1] - WWPT[i3])
                                WWPar=np.array(WWP)       
                                for i3, value in enumerate(WWP, start=num4):
                                    row3 = i3
                                    self.sheet.cell(row=row3, column=7).value = value
                                num4 = num4 + len(WGP)
                            else:
                                  pass

                            if len(WOPT) > 0:  # Check if WOPT array is not empty
                                WOP = [WOPT[0]]

                                for i in range(len(WOPT)-1): 
                                    WOP.append(WOPT[i+1] - WOPT[i])
  
                                for i, value in enumerate(WOP, start=num):
                                    row = i

                                    self.sheet.cell(row=row, column=5).value = value
                                    

                                    self.sheet.cell(row=row, column=1).value = f"{sim_well.name}"
                                    self.sheet.cell(row=row, column=2).value = f"{sim_well.name}"

                                    self.sheet.cell(row=row, column=4).value = step[row - num]

                                    # Add the code snippet at the third column
                                    summary_data_sampled = summary_case.resample_values("FOPT")
                                    t = summary_data_sampled.time_steps[row - num]  # Get the corresponding time step
                 
                                    self.sheet.cell(row=row, column=3).value = time.strftime("%d %b %Y", time.gmtime(t))
                                    
                                    
                                    
                                    # Combine values into a single SQL statement
                                    if sql_t=='ACTIVE':
                                       insert_query = f"INSERT INTO {cases[case_num-1].name}_PROD_WELL (WellName, ReportDate, Days, WellOil, WellGas, WellWater) VALUES (?, ?, ?, ?, ?,?)"

                                    # Create a list of tuples containing the values to insert
                                       values_to_insert = [(f"{sim_well.name}", time.strftime("%d %b %Y", time.gmtime(t)), step[row - num], value, val, va) for val, va in zip(WGPar, WWPar)]


                                       self.cursor.executemany(insert_query, values_to_insert)
                                       self.connection.commit()

                                num = num + len(WOP)
                            else:
                                pass
     
                     
                    for sim_well in sim_wells:

                            WWIT = np.array(summary_case.resample_values(f"WWIT:{sim_well.name}").values)
                            WGIT = np.array(summary_case.resample_values(f"WGIT:{sim_well.name}").values)

                            count = 0 
                            count1 = 0
                            count2 = 0
                            if len(WGIT) > 0 and len(WWIT) > 0:

                                    for value in WGIT:
                                        if value != 0:
                                            break
                                        count1 += 1    

                                    for value in WWIT:
                                        if value != 0:
                                            break
                                        count2 += 1

                                    if count1 < count2: 
                                        count=count1

                                    elif count2 < count1: 
                                        count=count2
                                    else: 
                                        count=count2
                                    WGI = [WGIT[0]]
                                    for j1 in range(len(WGIT)-1): 
                                        WGI.append(WGIT[j1+1] - WGIT[j1])
                                    WGI=WGI[count:]
                                    

                                    for j2, y in enumerate(WGI, start=num2):
                                        r1 = j2
                                        self.inj_sheet.cell(row=r1, column=6).value = y
                                         
                                        n_wgi.append(int(y))
                                            
                                        self.inj_sheet.cell(row=r1, column=4).value = step[r1+count- num2]
                                        n_step.append(step[r1+count- num2])
                                        self.inj_sheet.cell(row=r1, column=1).value = f"{sim_well.name}"
                                        self.inj_sheet.cell(row=r1, column=2).value = f"{sim_well.name}"
                                        
                                        
                                        
                                        n_well.append(f"{sim_well.name}")
                                        summary_data_sampled = summary_case.resample_values("TIME")
                                        t=summary_data_sampled.time_steps
                                        t=t[count:] 

                                        self.inj_sheet.cell(row=r1, column=3).value=time.strftime("%d %b %Y", time.gmtime(t[r1-num2]))
                                        n_time.append(time.strftime("%d %b %Y", time.gmtime(t[r1-num2])))

                                    num2 = num2 + len(WGI)

                                    WWI = [WWIT[0]]
                                    for j in range(len(WWIT)-1): 

                                        WWI.append(WWIT[j+1] - WWIT[j])
                                    WWI=WWI[count:] 
                                    

                                    for j5, W in enumerate(WWI, start=num1): 
                                            r = j5
                                            self.inj_sheet.cell(row=r, column=5).value = W
                                            
                                            n_wwi.append(int(W))
                                            
                                            self.inj_sheet.cell(row=r, column=4).value = step[r+count- num1]
                                            n_step.append(step[r+count- num1])

                                            self.inj_sheet.cell(row=r, column=1).value = f"{sim_well.name}"
                                            self.inj_sheet.cell(row=r, column=2).value = f"{sim_well.name}"
                                            
                                            n_well.append(f"{sim_well.name}")
                                            
                                            summary_data_sampled = summary_case.resample_values("TIME")
                                            t=summary_data_sampled.time_steps
                                            t=t[count:] 

                                            self.inj_sheet.cell(row=r, column=3).value=time.strftime("%d %b %Y", time.gmtime(t[r-num1]))
                                            n_time.append(time.strftime("%d %b %Y", time.gmtime(t[r-num1])))
                                    num1 = num1 + len(WWI)

                            elif len(WGIT)> 0 and len(WWIT)==0 :

                                        for value in WGIT:
                                            if value != 0:
                                                break
                                            count += 1
                                        WGI = [WGIT[0]]
                                        for j1 in range(len(WGIT)-1): 
                                            WGI.append(WGIT[j1+1] - WGIT[j1])
                                        WGI=WGI[count:]   
                                        
                                         
                                        for j2, y in enumerate(WGI, start=num2):
                                            r1 = j2
                                            self.inj_sheet.cell(row=r1, column=6).value = y
                                            
                                            n_wgi.append(int(y))
                                            
                                            self.inj_sheet.cell(row=r1, column=4).value = step[r1+count- num2]
                                            n_step.append(step[r1+count- num2])
                                            self.inj_sheet.cell(row=r1, column=1).value = f"{sim_well.name}"
                                            self.inj_sheet.cell(row=r1, column=2).value = f"{sim_well.name}"
                                            n_well.append(f"{sim_well.name}")
                                           
                                            t=summary_data_sampled.time_steps
                                            t=t[count:] 

                                            self.inj_sheet.cell(row=r1, column=3).value=time.strftime("%d %b %Y", time.gmtime(t[r1-num2]))
                                            n_time.append(time.strftime("%d %b %Y", time.gmtime(t[r1-num2])))
                                        num2 = num2 + len(WGI)

                            elif len(WWIT)> 0 and len(WGIT)==0:

                                        for value in WWIT:
                                            if value != 0:
                                                break
                                            count += 1
                                        WWI = [WWIT[0]]
                                        for j in range(len(WWIT)-1): 

                                            WWI.append(WWIT[j+1] - WWIT[j])
                                        WWI=WWI[count:] 
                                       

                                        for j5, W in enumerate(WWI, start=num1): 
                                            r = j5
                                            self.inj_sheet.cell(row=r, column=5).value = W
                                            
                                            n_wwi.append(int(W))
                                           
                                        
                                            

                                            self.inj_sheet.cell(row=r, column=4).value = step[r+count- num1]
                                            
                                            n_step.append(step[r+count- num1])

                                            self.inj_sheet.cell(row=r, column=1).value = f"{sim_well.name}"
                                            self.inj_sheet.cell(row=r, column=2).value = f"{sim_well.name}"
                                            
                                            n_well.append(f"{sim_well.name}")
                                            t=summary_data_sampled.time_steps
                                            t=t[count:] 

                                            self.inj_sheet.cell(row=r, column=3).value=time.strftime("%d %b %Y", time.gmtime(t[r-num1]))
                                            n_time.append(time.strftime("%d %b %Y", time.gmtime(t[r-num1])))
                                        num1 = num1 + len(WWI)
                                         
                                    
                            else:
                                   pass       
        
                           
                            
                            
                    if saving_path is  None:
                       saving_path=rf"{cases[case_num-1].name}_sim_reprt.xlsx"

                    self.workbook.save(saving_path)
                    
                    
                        
        
                    print(f"Finished exporting the data of {cases[case_num-1].name} case")
                    
                    
                    if sql_t=='ACTIVE':
                                       insert_queryu = f"INSERT INTO {cases[case_num-1].name}_INJ_WELLS (API,WellName, ReportDate,Days, MonthlyWater, MonthlyGasREAL) VALUES (?,?,?, ?, ?, ?)"

                                    # Create a list of tuples containing the values to insert
                                       values_to_insertu = [(n_w,n_w,n_t,n_d,n_wa,n_o) for n_w,n_t,n_d,n_wa,n_o in zip(n_well,n_time,n_step,n_wwi,n_wgi)]


                                       self.cursor.executemany(insert_queryu, values_to_insertu)
                                       self.connection.commit()

                    
        else: 
            print("Resinsight is not open ")
            
        
                  